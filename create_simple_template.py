#!/usr/bin/env python3
"""
Clappia용 단순화 Excel 템플릿 (시트별 개별 파일)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

def create_accounts_file():
    """구글 아이디 관리 파일"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"

    headers = ["account_id", "google_id", "memo", "active"]
    ws.append(headers)

    # 헤더 스타일
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # 샘플 데이터
    samples = [
        ["A01", "user1@gmail.com", "Main account", "TRUE"],
        ["A02", "user2@gmail.com", "Sub account", "TRUE"],
        ["A03", "user3@gmail.com", "Test account", "TRUE"],
    ]
    for row in samples:
        ws.append(row)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10

    wb.save("01_Accounts.xlsx")
    print("✅ 01_Accounts.xlsx 생성")

def create_services_file():
    """사이트 정보 파일"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Services"

    headers = ["service_id", "site_name", "url", "daily_limit", "reset_cycle"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    samples = [
        ["S01", "ChatGPT", "https://chat.openai.com", 50, "DAILY"],
        ["S02", "Gemini", "https://gemini.google.com", 60, "DAILY"],
        ["S03", "Claude", "https://claude.ai", 45, "DAILY"],
        ["S04", "Perplexity", "https://perplexity.ai", 5, "DAILY"],
    ]
    for row in samples:
        ws.append(row)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    wb.save("02_Services.xlsx")
    print("✅ 02_Services.xlsx 생성")

def create_mapping_file():
    """아이디-사이트 매핑 파일"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Account_Service"

    headers = ["account_id", "service_id", "subscribed"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    samples = [
        ["A01", "S01", "TRUE"],
        ["A01", "S02", "TRUE"],
        ["A01", "S03", "TRUE"],
        ["A02", "S01", "TRUE"],
        ["A02", "S04", "TRUE"],
        ["A03", "S02", "TRUE"],
        ["A03", "S03", "TRUE"],
    ]
    for row in samples:
        ws.append(row)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12

    wb.save("03_Account_Service.xlsx")
    print("✅ 03_Account_Service.xlsx 생성")

def create_daily_usage_file():
    """매일 사용량 기록 파일 (메인)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily_Usage"

    headers = ["date", "account_id", "service_id", "total_quota", "usage", "remaining", "status"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    # 최근 3일치 샘플
    today = datetime.now()
    samples = []

    for i in range(3):
        date = (today - timedelta(days=2-i)).strftime("%Y-%m-%d")
        samples.extend([
            [date, "A01", "S01", 50, 12, 38, "SAFE"],
            [date, "A01", "S02", 60, 35, 25, "WARNING"],
            [date, "A02", "S01", 50, 45, 5, "DANGER"],
        ])

    for row in samples:
        ws.append(row)

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 12

    wb.save("04_Daily_Usage.xlsx")
    print("✅ 04_Daily_Usage.xlsx 생성")

def create_all_in_one_simple():
    """올인원 단순 버전 (영문, 이모지 제거)"""
    wb = Workbook()

    # 1. Accounts
    ws1 = wb.active
    ws1.title = "Accounts"
    ws1.append(["account_id", "google_id", "memo", "active"])
    ws1.append(["A01", "user1@gmail.com", "Main", "TRUE"])
    ws1.append(["A02", "user2@gmail.com", "Sub", "TRUE"])

    # 2. Services
    ws2 = wb.create_sheet("Services")
    ws2.append(["service_id", "site_name", "daily_limit"])
    ws2.append(["S01", "ChatGPT", 50])
    ws2.append(["S02", "Gemini", 60])
    ws2.append(["S03", "Claude", 45])

    # 3. Daily_Usage (메인)
    ws3 = wb.create_sheet("Daily_Usage")
    ws3.append(["date", "account_id", "service_id", "usage"])
    ws3.append([datetime.now().strftime("%Y-%m-%d"), "A01", "S01", 10])
    ws3.append([datetime.now().strftime("%Y-%m-%d"), "A01", "S02", 15])

    wb.save("Simple_Template.xlsx")
    print("✅ Simple_Template.xlsx 생성 (최소 구조)")

if __name__ == "__main__":
    print("📦 Clappia용 Excel 파일 생성 중...\n")

    print("방법 1: 시트별 개별 파일")
    create_accounts_file()
    create_services_file()
    create_mapping_file()
    create_daily_usage_file()

    print("\n방법 2: 올인원 단순 버전")
    create_all_in_one_simple()

    print("\n✅ 완료!")
    print("\n📋 사용법:")
    print("1. Simple_Template.xlsx 먼저 시도 (가장 단순)")
    print("2. 안 되면 04_Daily_Usage.xlsx만 업로드")
    print("3. 각 시트를 개별 앱으로 만들고 싶으면 01~04 파일 각각 업로드")
