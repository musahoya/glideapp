#!/usr/bin/env python3
"""
Clappia 최적화 Excel 템플릿 생성 스크립트
구글 아이디 무료할당량 관리 앱용
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

def create_template():
    wb = Workbook()

    # ==================== 1. Accounts 시트 ====================
    ws_accounts = wb.active
    ws_accounts.title = "Accounts"

    # 헤더 (Clappia가 자동 인식)
    headers_accounts = ["account_id", "google_id", "메모", "활성"]
    ws_accounts.append(headers_accounts)

    # 헤더 스타일
    for cell in ws_accounts[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 샘플 데이터
    sample_accounts = [
        ["A01", "example1@gmail.com", "주력 계정", "TRUE"],
        ["A02", "example2@gmail.com", "서브 계정", "TRUE"],
        ["A03", "example3@gmail.com", "테스트용", "TRUE"],
    ]
    for row in sample_accounts:
        ws_accounts.append(row)

    # 열 너비 조정
    ws_accounts.column_dimensions['A'].width = 12
    ws_accounts.column_dimensions['B'].width = 25
    ws_accounts.column_dimensions['C'].width = 15
    ws_accounts.column_dimensions['D'].width = 10

    # ==================== 2. Services 시트 ====================
    ws_services = wb.create_sheet("Services")

    headers_services = ["service_id", "사이트명", "URL", "일일무료한도", "리셋주기"]
    ws_services.append(headers_services)

    # 헤더 스타일
    for cell in ws_services[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 샘플 데이터 (실제 서비스들)
    sample_services = [
        ["S01", "ChatGPT", "https://chat.openai.com", 50, "DAILY"],
        ["S02", "Gemini", "https://gemini.google.com", 60, "DAILY"],
        ["S03", "Claude", "https://claude.ai", 45, "DAILY"],
        ["S04", "Perplexity", "https://perplexity.ai", 5, "DAILY"],
    ]
    for row in sample_services:
        ws_services.append(row)

    # 열 너비 조정
    ws_services.column_dimensions['A'].width = 12
    ws_services.column_dimensions['B'].width = 15
    ws_services.column_dimensions['C'].width = 30
    ws_services.column_dimensions['D'].width = 15
    ws_services.column_dimensions['E'].width = 12

    # ==================== 3. Account_Service 시트 ====================
    ws_mapping = wb.create_sheet("Account_Service")

    headers_mapping = ["account_id", "service_id", "가입여부"]
    ws_mapping.append(headers_mapping)

    # 헤더 스타일
    for cell in ws_mapping[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 샘플 데이터 (가입 매핑)
    sample_mapping = [
        ["A01", "S01", "TRUE"],
        ["A01", "S02", "TRUE"],
        ["A01", "S03", "TRUE"],
        ["A02", "S01", "TRUE"],
        ["A02", "S04", "TRUE"],
        ["A03", "S02", "TRUE"],
        ["A03", "S03", "TRUE"],
    ]
    for row in sample_mapping:
        ws_mapping.append(row)

    # 열 너비 조정
    ws_mapping.column_dimensions['A'].width = 12
    ws_mapping.column_dimensions['B'].width = 12
    ws_mapping.column_dimensions['C'].width = 12

    # ==================== 4. Daily_Usage 시트 (메인) ====================
    ws_usage = wb.create_sheet("Daily_Usage")

    headers_usage = ["날짜", "account_id", "service_id", "총할당량", "사용량", "남은량", "상태"]
    ws_usage.append(headers_usage)

    # 헤더 스타일
    for cell in ws_usage[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 샘플 데이터 (최근 3일치)
    today = datetime.now()
    sample_usage = []

    for i in range(3):
        date = (today - timedelta(days=2-i)).strftime("%Y-%m-%d")
        sample_usage.extend([
            [date, "A01", "S01", 50, 12, 38, "안정"],
            [date, "A01", "S02", 60, 35, 25, "주의"],
            [date, "A02", "S01", 50, 45, 5, "위험"],
        ])

    for row in sample_usage:
        ws_usage.append(row)

    # 열 너비 조정
    ws_usage.column_dimensions['A'].width = 12
    ws_usage.column_dimensions['B'].width = 12
    ws_usage.column_dimensions['C'].width = 12
    ws_usage.column_dimensions['D'].width = 12
    ws_usage.column_dimensions['E'].width = 10
    ws_usage.column_dimensions['F'].width = 10
    ws_usage.column_dimensions['G'].width = 10

    # ==================== 5. 사용 가이드 시트 ====================
    ws_guide = wb.create_sheet("사용가이드")

    guide_content = [
        ["🚀 Clappia 앱 만들기 가이드"],
        [""],
        ["📋 이 엑셀 파일 구조"],
        ["1. Accounts - 구글 아이디 관리"],
        ["2. Services - 가입한 사이트 정보"],
        ["3. Account_Service - 어떤 아이디가 어떤 사이트에 가입했는지"],
        ["4. Daily_Usage - 매일 사용량 기록 (⭐ 메인)"],
        [""],
        ["✅ Clappia로 앱 만드는 3단계"],
        [""],
        ["1️⃣ Clappia 가입"],
        ["   → https://www.clappia.com 접속"],
        ["   → Google 계정으로 무료 가입"],
        [""],
        ["2️⃣ 앱 생성"],
        ["   → 왼쪽 메뉴 'More' → 'Create App from Excel'"],
        ["   → 이 파일 업로드"],
        ["   → 각 시트별로 앱 생성됨"],
        [""],
        ["3️⃣ 앱 커스터마이징"],
        ["   → Daily_Usage 앱을 메인으로 설정"],
        ["   → account_id, service_id를 Dropdown으로 변경"],
        ["   → 남은량 기준 색상 조건부 서식 추가"],
        [""],
        ["💡 핵심 팁"],
        [""],
        ["✔ 구글 아이디 추가 = Accounts 시트에 행 추가"],
        ["✔ 사이트 추가 = Services 시트에 행 추가"],
        ["✔ 매일 Daily_Usage만 입력하면 됨"],
        ["✔ Clappia 무료 플랜: 무제한 앱 + 100명 사용자"],
        [""],
        ["🔗 다음 단계: CLAPPIA_GUIDE.md 파일 참고"],
    ]

    for row in guide_content:
        ws_guide.append(row)

    # 제목 스타일
    ws_guide['A1'].font = Font(bold=True, size=16, color="C00000")
    ws_guide.column_dimensions['A'].width = 60

    # 파일 저장
    filename = "구글아이디_무료할당량_관리_Clappia템플릿.xlsx"
    wb.save(filename)
    print(f"✅ Excel 템플릿 생성 완료: {filename}")
    print(f"📊 총 {len(wb.sheetnames)}개 시트:")
    for sheet in wb.sheetnames:
        print(f"   - {sheet}")

if __name__ == "__main__":
    create_template()
